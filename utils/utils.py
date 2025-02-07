from openpyxl import Workbook, load_workbook

def save_excel(archivo_excel: str, data: list, sheet_name:str, column=1):    
    try:
        # Intenta cargar el archivo existente
        libro = load_workbook(archivo_excel)
    except FileNotFoundError:
        # Si el archivo no existe, crea uno nuevo
        libro = Workbook()

    # Selecciona la hoja llamada "nombres" o créala si no existe
    if sheet_name in libro.sheetnames:
        hoja = libro[sheet_name]
    else:
        hoja = libro.create_sheet(sheet_name)

    # Encuentra la primera fila vacía en la columna A
    fila_vacia = hoja.max_row + 1

    # Agrega los nuevos nombres a la hoja
    for i, nombre in enumerate(data, start=fila_vacia):
        hoja.cell(row=i, column=column, value=nombre)

    # Guarda los cambios en el archivo Excel
    libro.save(archivo_excel)

    print(f"Se han agregado {len(data)} {sheet_name} al archivo.")